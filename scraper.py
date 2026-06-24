import asyncio
import csv
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.async_api import async_playwright

from municipios_sp import MUNICIPIOS_SP
from supabase_history import carregar_lojas as carregar_historico_supabase
from supabase_history import salvar_lojas as salvar_historico_supabase
from supabase_history import supabase_ativo

REGIOES_DISPONIVEIS = MUNICIPIOS_SP

MAX_WORKERS = 3
CAMPOS = [
    "regiao", "nome", "telefone", "whatsapp", "endereco",
    "categoria", "horario", "site", "nota", "url"
]
HISTORICO_ARQUIVO = "lojas_ja_coletadas.csv"
TERMOS_MOVEIS = [
    "movel", "moveis", "móveis", "mobiliario", "mobiliário",
    "decoracao", "decoração", "colchoes", "colchões", "estofado",
    "estofados", "sofa", "sofá", "planejados", "marcenaria",
    "cama", "mesa", "cadeira", "armario", "armário", "guarda roupa",
    "guarda-roupa", "rack", "decor", "casa"
]
TERMOS_BLOQUEADOS = [
    "supermercado", "mercado", "hipermercado", "mercearia",
    "departamento", "eletrodomestico", "eletrodoméstico", "eletronico",
    "eletrônico", "farmacia", "farmácia", "padaria", "restaurante",
    "lanchonete", "shopping", "posto de gasolina", "autopecas",
    "autopeças", "material de construcao", "material de construção",
    "conveniencia", "conveniência", "variedades", "roupa", "calcados",
    "calçados", "pet shop", "academia", "oficina", "hotel"
]

ProgressCallback = Optional[Callable[[Dict], None]]


def formatar_telefone(numero: str) -> str:
    if not numero:
        return ""

    digits = "".join(c for c in numero if c.isdigit())

    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]

    if digits.startswith("0") and len(digits) in (11, 12):
        digits = digits[1:]

    if len(digits) == 11:
        ddd = digits[:2]
        resto = digits[2:]
        return f"({ddd}) {resto[0]} {resto[1:5]}-{resto[5:]}"

    if len(digits) == 10:
        ddd = digits[:2]
        resto = digits[2:]
        return f"({ddd}) {resto[:4]}-{resto[4:]}"

    return numero


def limpar_texto(texto: str) -> str:
    if not texto:
        return ""

    try:
        return texto.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore").strip()
    except Exception:
        return str(texto).strip()


def log_erro(log_path: Path, mensagem: str) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {mensagem}\n")


def chave_texto(valor: str) -> str:
    valor = limpar_texto(valor or "").lower()
    valor = re.sub(r"https?://(www\.)?", "", valor)
    valor = re.sub(r"[^a-z0-9]+", " ", valor)
    return re.sub(r"\s+", " ", valor).strip()


def texto_busca(valor: str) -> str:
    valor = limpar_texto(valor or "").lower()
    valor = valor.replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
    valor = valor.replace("é", "e").replace("ê", "e")
    valor = valor.replace("í", "i")
    valor = valor.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    valor = valor.replace("ú", "u").replace("ü", "u")
    valor = valor.replace("ç", "c")
    return valor


def parece_loja_de_moveis(loja: Dict) -> bool:
    texto = texto_busca(
        " ".join(
            [
                loja.get("nome", ""),
                loja.get("categoria", ""),
                loja.get("site", ""),
                loja.get("url", ""),
            ]
        )
    )

    bloqueado = any(texto_busca(termo) in texto for termo in TERMOS_BLOQUEADOS)
    permitido = any(texto_busca(termo) in texto for termo in TERMOS_MOVEIS)

    return permitido and not bloqueado


def chave_telefone(valor: str) -> str:
    digits = "".join(c for c in (valor or "") if c.isdigit())
    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]
    if digits.startswith("0") and len(digits) in (11, 12):
        digits = digits[1:]
    return digits


def chaves_loja(loja: Dict) -> Set[str]:
    chaves: Set[str] = set()

    url = chave_texto(loja.get("url", ""))
    if url:
        chaves.add(f"url:{url}")

    telefone = chave_telefone(loja.get("telefone", ""))
    if telefone:
        chaves.add(f"tel:{telefone}")

    whatsapp = chave_telefone(loja.get("whatsapp", ""))
    if whatsapp:
        chaves.add(f"tel:{whatsapp}")

    site = chave_texto(loja.get("site", ""))
    if site:
        chaves.add(f"site:{site}")

    nome = chave_texto(loja.get("nome", ""))
    endereco = chave_texto(loja.get("endereco", ""))
    if nome and endereco:
        chaves.add(f"nome_endereco:{nome}|{endereco}")
    elif nome:
        chaves.add(f"nome:{nome}")

    return chaves


def carregar_csv_lojas(path: Path) -> List[Dict]:
    if not path.exists() or path.stat().st_size == 0:
        return []

    try:
        with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader if row.get("nome") or row.get("url")]
    except Exception:
        return []


def carregar_historico(outputs_root: Path) -> Tuple[List[Dict], Set[str]]:
    if supabase_ativo():
        lojas: List[Dict] = carregar_historico_supabase()
    else:
        historico_path = outputs_root / HISTORICO_ARQUIVO
        lojas = carregar_csv_lojas(historico_path)

        for csv_path in outputs_root.glob("*/*.csv"):
            lojas.extend(carregar_csv_lojas(csv_path))

    chaves: Set[str] = set()
    unicas: List[Dict] = []
    for loja in lojas:
        if not parece_loja_de_moveis(loja):
            continue

        loja_chaves = chaves_loja(loja)
        if not loja_chaves or chaves.intersection(loja_chaves):
            continue
        chaves.update(loja_chaves)
        unicas.append({campo: loja.get(campo, "") for campo in CAMPOS})

    return unicas, chaves


def salvar_historico(outputs_root: Path, lojas: List[Dict]) -> None:
    if supabase_ativo():
        salvar_historico_supabase(lojas)
        return

    historico_path = outputs_root / HISTORICO_ARQUIVO
    salvar_csv(lojas, historico_path)


def filtrar_lojas_novas(dados: List[Dict], chaves_historico: Set[str]) -> Tuple[List[Dict], int]:
    novas: List[Dict] = []
    chaves_vistas = set(chaves_historico)
    duplicadas = 0

    for loja in dados:
        loja_chaves = chaves_loja(loja)
        if loja_chaves and chaves_vistas.intersection(loja_chaves):
            duplicadas += 1
            continue

        novas.append(loja)
        chaves_vistas.update(loja_chaves)

    return novas, duplicadas


async def buscar_links_regiao(page, regiao: str, max_lojas: int, log_path: Path) -> List[str]:
    busca = f"lojas+de+moveis+{regiao.replace(' ', '+')}+SP"
    url = f"https://www.google.com/maps/search/{busca}"

    for tentativa in range(3):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(3)
            break
        except Exception as e:
            if tentativa == 2:
                log_erro(log_path, f"Falha ao carregar regiÃ£o {regiao}: {e}")
                return []
            await asyncio.sleep(3)

    conteudo = await page.content()
    if "captcha" in conteudo.lower():
        log_erro(log_path, f"Captcha detectado em {regiao}.")
        return []

    for _ in range(max(5, max_lojas // 4)):
        try:
            lista = page.locator('div[role="feed"]')
            if await lista.count() > 0:
                await lista.first.evaluate("el => el.scrollBy(0, 800)")
            else:
                await page.evaluate("window.scrollBy(0, 800)")
        except Exception:
            await page.evaluate("window.scrollBy(0, 800)")
        await asyncio.sleep(0.8)

    links = await page.locator('a[href*="/maps/place/"]').all()
    hrefs = []
    for link in links:
        href = await link.get_attribute("href")
        if href and href not in hrefs:
            hrefs.append(href)
        if len(hrefs) >= max_lojas:
            break

    return hrefs


async def extrair_detalhes(page, url: str, regiao: str, log_path: Path) -> Optional[Dict]:
    for tentativa in range(3):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(2)

            nome = ""
            try:
                nome = await page.locator("h1").first.inner_text(timeout=3000)
            except Exception:
                pass

            if not nome:
                return None

            telefone = ""
            try:
                tel_el = page.locator('button[data-item-id^="phone"]')
                if await tel_el.count() > 0:
                    telefone = await tel_el.first.get_attribute("data-item-id")
                    telefone = telefone.replace("phone:tel:", "").strip()
            except Exception:
                pass

            whatsapp = ""
            try:
                wa_el = page.locator('a[href*="wa.me"]')
                if await wa_el.count() > 0:
                    wa_href = await wa_el.first.get_attribute("href")
                    whatsapp = wa_href.replace("https://wa.me/", "").split("?")[0]
            except Exception:
                pass

            endereco = ""
            try:
                end_el = page.locator('button[data-item-id="address"]')
                if await end_el.count() > 0:
                    raw = await end_el.first.get_attribute("aria-label") or await end_el.first.inner_text(timeout=3000)
                    endereco = raw.replace("EndereÃ§o: ", "").replace("Address: ", "").strip()
            except Exception:
                pass

            site = ""
            try:
                site_el = page.locator('a[data-item-id="authority"]')
                if await site_el.count() > 0:
                    site = await site_el.first.get_attribute("href")
            except Exception:
                pass

            nota = ""
            seletores_nota = [
                'div.F7nice span[aria-hidden="true"]',
                "span.ceNzKf",
                'div[jsaction*="pane.rating"]',
                'span[aria-hidden="true"]',
            ]
            for seletor in seletores_nota:
                try:
                    els = await page.locator(seletor).all()
                    for el in els:
                        texto = await el.inner_text(timeout=1500)
                        match = re.search(r"\b[1-5][.,]\d\b", texto)
                        if match:
                            nota = match.group().replace(",", ".")
                            break
                    if nota:
                        break
                except Exception:
                    continue

            categoria = ""
            try:
                cat_el = page.locator('button[jsaction*="category"]').first
                categoria = await cat_el.inner_text(timeout=2000)
            except Exception:
                pass

            horario = ""
            try:
                hor_el = page.locator("div[data-hide-tooltip-on-mouse-move]").first
                horario = await hor_el.inner_text(timeout=2000)
                horario = horario.split("\n")[0].strip()
            except Exception:
                pass

            return {
                "regiao": regiao,
                "nome": limpar_texto(nome),
                "telefone": telefone,
                "whatsapp": whatsapp,
                "endereco": limpar_texto(endereco),
                "categoria": limpar_texto(categoria),
                "horario": limpar_texto(horario),
                "site": site or "",
                "nota": nota,
                "url": url,
            }

        except Exception as e:
            if tentativa == 2:
                log_erro(log_path, f"Falha ao extrair {url}: {e}")
            await asyncio.sleep(3)

    return None


async def worker(browser, fila: List[Tuple[str, str]], resultados: List[Dict], lock: asyncio.Lock, log_path: Path, progress_cb: ProgressCallback = None):
    page = await browser.new_page()
    while True:
        async with lock:
            if not fila:
                break
            url, regiao = fila.pop(0)

        dados = await extrair_detalhes(page, url, regiao, log_path)

        async with lock:
            if dados:
                resultados.append(dados)
            if progress_cb:
                progress_cb({"type": "item_done", "current": len(resultados)})

    await page.close()


def salvar_csv(dados: List[Dict], arquivo: Path) -> None:
    arquivo.parent.mkdir(parents=True, exist_ok=True)
    with arquivo.open("w", newline="", encoding="utf-8-sig", errors="replace") as f:
        writer = csv.DictWriter(f, fieldnames=CAMPOS)
        writer.writeheader()
        writer.writerows(dados)


def salvar_excel(dados: List[Dict], arquivo: Path) -> None:
    arquivo.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lojas de MÃ³veis SP"

    cabecalhos = ["RegiÃ£o", "Nome", "Telefone", "WhatsApp", "EndereÃ§o", "Categoria", "HorÃ¡rio", "Site", "Nota", "URL"]
    chaves = CAMPOS

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=cab)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, loja in enumerate(dados, 2):
        for col_idx, chave in enumerate(chaves, 1):
            ws.cell(row=row_idx, column=col_idx, value=loja.get(chave, ""))

    larguras = [15, 35, 18, 18, 45, 25, 20, 35, 8, 50]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura

    ws.freeze_panes = "A2"
    wb.save(arquivo)


def tratar_resultados(resultados: List[Dict]) -> List[Dict]:
    vistos_nome = set()
    unicos = []

    for d in resultados:
        for campo in ["nome", "endereco", "categoria", "horario", "site", "url", "regiao"]:
            d[campo] = limpar_texto(d.get(campo, ""))

        if not parece_loja_de_moveis(d):
            continue

        chave = d.get("nome", "").lower().strip()
        if not chave or chave in vistos_nome:
            continue

        vistos_nome.add(chave)
        d["telefone"] = formatar_telefone(d.get("telefone", ""))
        d["whatsapp"] = formatar_telefone(d.get("whatsapp", ""))

        unicos.append(d)

    return unicos


def montar_resumo(dados: List[Dict], regioes: List[str]) -> Dict:
    total = len(dados)
    com_telefone = sum(1 for d in dados if d.get("telefone"))
    com_whatsapp = sum(1 for d in dados if d.get("whatsapp"))
    com_site = sum(1 for d in dados if d.get("site"))

    por_regiao = {regiao: sum(1 for d in dados if d.get("regiao") == regiao) for regiao in regioes}

    return {
        "total": total,
        "com_telefone": com_telefone,
        "com_whatsapp": com_whatsapp,
        "com_site": com_site,
        "por_regiao": por_regiao,
    }


async def executar_varredura(regioes: List[str], max_lojas: int, output_dir: Path, headless: bool = True, progress_cb: ProgressCallback = None) -> Dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs_root = output_dir.parent
    historico_lojas, historico_chaves = carregar_historico(outputs_root)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    data_arquivo = datetime.now().strftime("%d-%m-%Y")
    nome_base = f"Varredura-SP {data_arquivo}"
    log_path = output_dir / f"erros_{timestamp}.log"
    csv_path = output_dir / f"{nome_base}.csv"
    excel_path = output_dir / f"{nome_base}.xlsx"
    duplicadas_por_url = 0
    duplicadas_por_dados = 0
    
    if progress_cb:
        progress_cb({"type": "message", "message": f"Preparando busca. Historico atual: {len(historico_lojas)} lojas."})

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        page_init = await browser.new_page()

        await page_init.goto("https://www.google.com/maps", wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(3)

        try:
            botao = page_init.locator('button:has-text("Aceitar tudo")')
            if await botao.count() > 0:
                await botao.first.click()
                await asyncio.sleep(1)
        except Exception:
            pass

        todos_links: List[Tuple[str, str]] = []
        for regiao in regioes:
            if progress_cb:
                progress_cb({"type": "message", "message": f"Coletando links em {regiao}..."})
            links = await buscar_links_regiao(page_init, regiao, max_lojas, log_path)
            todos_links.extend((link, regiao) for link in links)

        await page_init.close()

        vistos_url = set()
        fila: List[Tuple[str, str]] = []
        for url, regiao in todos_links:
            chave_url = f"url:{chave_texto(url)}"
            if url in vistos_url:
                continue
            vistos_url.add(url)

            if chave_url in historico_chaves:
                duplicadas_por_url += 1
                continue

            fila.append((url, regiao))

        if progress_cb:
            mensagem = f"Extraindo detalhes de {len(fila)} lojas novas..."
            if duplicadas_por_url:
                mensagem += f" {duplicadas_por_url} links repetidos foram ignorados."
            progress_cb({"type": "total", "total": len(fila), "message": mensagem})

        resultados: List[Dict] = []
        lock = asyncio.Lock()
        workers = [
            worker(browser, fila, resultados, lock, log_path, progress_cb)
            for _ in range(min(MAX_WORKERS, max(1, len(fila))))
        ]
        if workers:
            await asyncio.gather(*workers)

        await browser.close()

    unicos = tratar_resultados(resultados)
    novos, duplicadas_por_dados = filtrar_lojas_novas(unicos, historico_chaves)
    historico_atualizado = historico_lojas + novos

    salvar_csv(novos, csv_path)
    salvar_excel(novos, excel_path)
    salvar_historico(outputs_root, historico_atualizado)
    resumo = montar_resumo(novos, regioes)
    duplicadas_total = duplicadas_por_url + duplicadas_por_dados

    return {
        "dados": novos,
        "resumo": resumo,
        "csv_path": str(csv_path),
        "excel_path": str(excel_path),
        "log_path": str(log_path) if log_path.exists() else "",
        "duplicadas_ignoradas": duplicadas_total,
        "historico_total": len(historico_atualizado),
    }
